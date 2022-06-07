import datetime
import openpyxl
from datetime import time

class WrongDate():
    def __init__(self, date, classroom):
        self.date = date
        self.classroom = classroom

class WrongSubject():
    def __init__(self, name, classroom, cell_address):
        self.name = name
        self.classroom = classroom
        self.cell_address = cell_address

#settings.
date_column = 2
hr_row = 4
min_row = 5
#Check the difference between date (in days).
check_difference_date = 270

def read_excel_file(excel_file):

    return_list = []

    wb = openpyxl.load_workbook(filename=excel_file, data_only=True)

    for active_sheet in wb.sheetnames:
        sheet = wb[active_sheet]
        for merged in sheet.merged_cells:
            start = merged.start_cell
            row = start.row

            #lookingo for date cells.
            date_cell = sheet.cell(row=row, column=date_column)
            return_date = {'classroom': None, 'date': None, 'name': None, 'start_time': None, 'end_time': None}
            if isinstance(start.value, str) and date_cell.value != None:
                if check_many_date(sheet, merged):
                    date_list = []
                    difference = merged.max_row - merged.min_row
                    jump = int(difference / 2)
                    looking_row = merged.min_row
                    for number in range(jump+1):
                        cell = sheet.cell(row=looking_row, column=date_column)
                        date_list.append(cell)
                        looking_row += 2
                    for date in date_list:
                        return_date['classroom'] = active_sheet
                        return_date['name'] = start.value
                        return_date['date'] = date.value
                        return_date['start_time'] = create_start_time(merged)
                        return_date['end_time'] = create_end_time(merged)
                        return_list.append(return_date)

                else:

                    return_date['classroom'] = active_sheet
                    return_date['name'] = start.value
                    return_date['date'] = date_cell.value
                    return_date['start_time'] = create_start_time(merged)
                    return_date['end_time'] = create_end_time(merged)
                    return_list.append(return_date)

    return return_list



def check_many_date(sheet, merged):
    """"Check whether the classes are connected in more than one date. Returns all dates"""
    date_cell1 = sheet.cell(row=merged.max_row, column=date_column)
    date_cell2 = sheet.cell(row=merged.min_row, column=date_column)
    if date_cell1 != date_cell2:
        many_date = True
    else:
        many_date = False
    return many_date

def create_start_time(merged):
    minute = None
    #Minutes.
    minute_time = (merged.min_col - date_column) % 4
    if minute_time == 1:
        minute = 0
    if minute_time == 2:
        minute = 15
    if minute_time == 3:
        minute = 30
    if minute_time == 0:
        minute = 45
    #Hours.
    hour_time = int((merged.min_col - date_column-1) / 4)
    hour = hour_time + 7
    to_time = time(hour, minute, 00)
    return to_time

def create_end_time(merged):
    minute = None
    #Minutes.
    minute_time = (merged.max_col - date_column) % 4
    if minute_time == 0:
        minute = 0
    if minute_time == 1:
        minute = 15
    if minute_time == 2:
        minute = 30
    if minute_time == 3:
        minute = 45
    #Hours.
    hour_time = int((merged.max_col - date_column) / 4)
    hour = hour_time + 7
    to_time = time(hour, minute, 00)
    return to_time


def check_mistake(wb):
    """"Check if cells are merged with good date"""
    list = []
    for active_sheet in wb.sheetnames:
        sheet = wb[active_sheet]
        for merged in sheet.merged_cells:
            start = merged.start_cell
            #Check if merged is part of plan.
            if isinstance(start.value, str) and start.row > min_row and start.column > date_column:
                if start.row % 2 != 0:
                    ws = WrongSubject(start.value, active_sheet, start.coordinate)
                    list.append(ws)
    return list


def check_date(wb):
    """"Check for any typo in date value"""
    date_list = []
    today = datetime.datetime.now()
    difference = datetime.timedelta(days=check_difference_date)
    for active_sheet in wb.sheetnames:
        sheet = wb[active_sheet]
        for merged in sheet.merged_cells:
            start = merged.start_cell
            # Check if merged is datetime class.
            if isinstance(start.value, datetime.datetime) and start.value not in date_list\
                    and (start.value.date() > today.date() + difference or start.value.date() < today.date() - difference):
                wd = WrongDate(start.value, active_sheet)
                date_list.append(wd)

    for x in date_list:
        print(x)

    return date_list


def get_classrooms(excel_file):
    """Create the list with classroom"""
    return_list = []

    wb = openpyxl.load_workbook(filename=excel_file, data_only=True)

    for active_sheet in wb.sheetnames:
        return_list.append(active_sheet)

    return return_list